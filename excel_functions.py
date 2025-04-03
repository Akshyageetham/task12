


from openpyxl import load_workbook


class AkshyaExcelReader:


   def __init__(self, excel_file, sheet_name):
       self.file = excel_file
       self.sheet = sheet_name


   """Fetch the Row Count from Excel File"""
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   """Fetch the Column Count from Excel File"""
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   """Read the data from Excel File"""
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       data = sheet.cell(row=row_number, column=column_number).value
       return data


   """Write the data into Excel File"""
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)